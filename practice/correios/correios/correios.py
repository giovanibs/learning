"""Módulo com métodos para realizar o Exercício Excel Correios. Faz a validação da demanda,
executa a busca no site dos correios e, finalmente, grava os dados em Excel."""
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import correios.constants as const
from time import sleep, time
import pandas as pd


class Correios(webdriver.Chrome):
    """Cria novo driver e possui métodos para realizar a validação da demanda,
        executar a busca no site dos correios e, finalmente, gravar os dados em Excel."""

    def __init__(self, driver_path=const.DRIVER_PATH, teardown=True):
        """Cria nova instância da classe Correios. Adiciona a pasta com 
        drivers no PATH (variáveis de ambiente)

        IMPORTANTE: comentar a linha no método __init__ caso não seja necessário adicionar
        o caminho dos drivers no PATH (variáveis de ambiente).
        
        Parâmetros
        ----------
            driver_path (str): Caminho onde estão salvos os drivers dos navegadores (pode ser alterado nas contantes)
            
            teardown (bool): Flag utilizada no método __exit__
        """
        self.driver_path = driver_path
        self.teardown = teardown

        # COMENTAR A LINHA ABAIXO SE NÃO PRECISAR ALTERAR O PATH NAS VARIÁVEIS DE AMBIENTE
        os.environ["PATH"] += os.pathsep + self.driver_path

        # cria instância da classe webdriver.Chrome para ter acesso a todos os seus métodos
        super(Correios, self).__init__()

        self.implicitly_wait(5)

    def __exit__(self, *args) -> None:
        """Fecha a webpage caso o atributo `teardown` da instância seja `True`.
        """
        if self.teardown:
            self.quit()

    def create_results_dir(self):
        """Cria os diretórios para demandas executadas e demandas recusadas (se necessário)."""
        try:
            for dir in [const.DENIED, const.COMPLETED]:
                new_dir = os.path.join(os.getcwd(), const.DATA_PATH, dir)
                os.makedirs(new_dir, exist_ok=True)
        except:
            print("Aconteceu algo de errado ao criar as pastas")

    def land_first_page(self):
        """  Abre nova página da web com a URL Base """
        self.get(const.BASE_URL)

    def error_msg(self, idx, msg):
        """Retorna a mensagem de erro para uma pesquisa.
        
        Parâmetros
        ----------
            idx (int): Linha Excel Entrada.
            
            msg (str): Código da mensagem de erro.
        
        Retorna
        ----------
        Um DataFrame com o Linha de Excel Entrada | Mensagem de Erro
        """
        resultados_msg = pd.DataFrame(
            [[idx, const.MSG_STATUS[msg], const.MSG_STATUS[msg]]],
            columns=[
                const.RESULTS_HEADER[0],  # Linha Excel Entrada
                const.RESULTS_HEADER[1],  # Critério de busca
                const.RESULTS_HEADER[7],  # Mensagem
            ],
        )
        return resultados_msg

    def execute_search(self, search_crit, search_param):
        """Executa o preenchimento dos dados no website e clica no botão de pesquisa.
        
        Parâmetros
        ----------
            search_crit (str): Critério de busca
            
            search_param (str): Parâmetro de busca dos campos CEP ou Nome
            
        Retorna
        ----------
            1 caso a pesquisa tenha resultados.

            0 caso a pesquisa não tenha resultados.
        """
        try:
            self.new_search(search_crit)
            search_input = self.find_element_by_id(
                const.CRITERIA["SEARCH_BTN_ID"][search_crit]
            )
            self.paste_keys(search_param, search_input)
            search_button = self.find_element_by_id("btn_pesquisar")
            search_button.click()
            # próximos passos: tirar esses sleeps()
            sleep(2)
            search_status = self.find_element_by_id("mensagem-resultado").text

            if search_status == const.MSG_STATUS["No_results"]:
                return 0
            # Como a busca é realizada em outra webpage, temos mensagem
            # de log diferente para cada ('Nome' ou 'CEP')
            elif search_status == const.CRITERIA["SEARCH_LOG"][search_crit]:
                return 1
        except:
            # próximos passos: levantar os possíveis erros para tratar
            print("Ocorreu algum erro na página.")
            return 0

    def get_results(self, file_src):
        """Consolida todos o procedimentos da classe para a automação:
        
        1) Faz a validação do arquivo `file_src` (se recusar a demanda, pula para o passo 6).
        
        2) Executa a pesquisa no site para cada linha de entrada.
        
        3) Obtém os dados da tabela gerada pelo site.

        4) Consolida os resultados de todas as linhas de entrada (se não houver resultados,
        marca a demanda como recusada e pula para o passo 6)
        
        5) Grava os dados no Excel de origem e salva no mesmo arquivo.
        
        6) Retorna o caminho do arquivo de saída.
        
        Parâmetros
        ----------
            file_src (str): Caminho do arquivo de entrada.
            
        Retorna
        ----------
            Tuple com 2 valores:
            1) Novo path do arquivo já com a pasta de destino de acordo com a validação da demanda.
            A pasta de destino pode ser definida manualmente nas constantes (costants.py) COMPLETED e DENIED.
            2) Status do job (completed | denied)
        """
        print("Validando o arquivo de demanda...")
        wb = load_workbook(file_src)
        search_table = self.validate_demand(wb)
        search_summary = pd.DataFrame(columns=const.SUMMARY_HEADER)
        self.create_results_dir()

        if not search_table:
            print("Demanda recusada. Movendo arquivo...")
            wb.close()
            return (self.new_file_path(file_src, const.DENIED), "denied")

        else:
            print("Demanda aceita. Procedendo para a busca...\n\n")

            # Dataframe que conterá todos os resultados obtidos na pesquisa
            resultados = pd.DataFrame(columns=const.RESULTS_HEADER)

            # Loop para cada linha de dados no Excel capturado
            for idx, row in enumerate(search_table, start=1):
                print(f"\n... Linha de dados número #{idx}...")

                # Valida a pesquisa
                search_msg, search_crit, search_param = self.validate_search(row)

                # SE obtiver sucesso na validação
                if search_msg[0] == "S":
                    print("Extraindo dados da tabela...")
                    records = self.get_table()
                    nan_filler = self.nan_filler(
                        idx, search_crit, search_param, const.MSG_STATUS[search_msg]
                    )

                    new_records = pd.concat(
                        [pd.DataFrame(columns=const.RESULTS_HEADER), records]
                    ).fillna(nan_filler)
                    resultados = pd.concat([resultados, new_records])

                    # Atualiza informações para o resumo
                    new_search_summary = self.new_search_summary(idx, row, search_msg)
                    search_summary = pd.concat([search_summary, new_search_summary])
                else:
                    print("Sem resultados")
                    resultados_msg = self.error_msg(idx, search_msg)
                    resultados = pd.concat([resultados, resultados_msg])  # .fillna('')

                    # Atualiza informações para o resumo
                    new_search_summary = self.new_search_summary(idx, row, search_msg)
                    search_summary = pd.concat([search_summary, new_search_summary])

            # caso não haja resultados obtidos para todas as linhas de demanda,
            # tomando como base o campo 'CEP'
            # mover Excel para pasta de demanda recusada
            if not resultados[const.RESULTS_HEADER[6]].count():
                return (self.new_file_path(file_src, const.DENIED), "denied")

            ws2 = wb.create_sheet(title="Resultados")
            for r in dataframe_to_rows(resultados, index=False, header=True):
                ws2.append(r)

            ws3 = wb.create_sheet(title="Resumo")
            for r in dataframe_to_rows(search_summary, index=False, header=True):
                ws3.append(r)

            wb.save(file_src)
        wb.close()
        return (self.new_file_path(file_src, const.COMPLETED), "completed")

    def get_table(self):
        """ Lê a tabela da página HTML resultante da pesquisa e grava os dados em um objeto
        'pandas.DataFrame'.

        Retorna
        ----------
           Dataframe com os dados obtidos da tabela na página com os resultados da pesquisa.
        """
        # variável para fazer um entre o número de resultados apontado no site e o obtido
        records_size = int(self.find_element_by_id("navegacao-total").text.split()[-1])
        page_records = pd.DataFrame()
        next_page = 1

        while next_page:
            sleep(2)
            page_records = page_records.append(pd.read_html(self.page_source))

            # verifica se o botão 'Próximo' está visível na página
            try:
                next_page = self.find_element_by_css_selector(
                    'a[class="botao proximo"]'
                )
            except:
                try:
                    next_page = self.find_element_by_css_selector(
                        'a[class="botao proximo esconde"]'
                    )
                except:
                    # Sinalizar para tratar caso haja alguma necessidade
                    print("****** Algo deu errado...")
                break

            next_page.click()

        records_rows, records_cols = page_records.shape
        # Check se pegou todos os dados
        if records_size != records_rows:
            print(
                f"Operação falhou.\nResultados na busca: {records_size}.\nRegistros obtidos: {records_rows}."
            )

        return page_records

    def nan_filler(self, idx, search_crit, search_param, search_msg):
        """Retorna um dict para ser utilizado no método `fillna`,
        que preenche os campos vazios de acordo com o valor apontado
        no nome da coluna.
        
        Parâmetros
        ----------
            idx (int): Linha Excel Entrada 
            
            search_crit (str): Critério de busca utilizado
            
            search_param (str): Parâmetro de busca dos campos CEP ou Nome
            
            search_msg (str): Mensagem de status da pesquisa

        Retorna
        ----------
        Retorna um dict para ser utilizado no método `fillna`,
        que preenche os campos vazios de acordo com o valor apontado
        no nome da coluna.
        
        """
        return {
            const.RESULTS_HEADER[0]: idx,
            const.RESULTS_HEADER[1]: search_crit,
            const.RESULTS_HEADER[2]: search_param,
            const.RESULTS_HEADER[7]: search_msg,
        }

    def new_file_path(self, file_src, job_status):
        """Gera o caminho/nome do arquivo de saída do job,
        adicionando uma timestamp ao nome do arquivo e
        definindo a pasta de destino conforme a
        validação da demanda.
        
        Parâmetros
        ----------
            file_src (str):     Caminho do arquivo original
            
            job_status (str):   Mensagem de validação da demanda
            
        Retorna
        ----------
            Novo path do arquivo já com a pasta de destino de acordo com a validação da demanda.
        """
        # Timestamp
        ts = time()
        # caminho da pasta, nome do arquivo
        root_path, base_name = os.path.split(file_src)
        # novo nome com timestamp
        new_name = base_name.replace(".xlsx", f"_{ts}.xlsx")
        dest = os.path.join(root_path, job_status, new_name)
        return dest

    def new_search(self, search_crit):
        """Aponta o navegador para a página de pesquisa conforme critério: CEP ou Nome.

        A página de pesquisa por CEP ajuda a mitigar a necessidade de filtrar o número
        do CEP encontrado no campo Logradouro/Localidade (eg: CEP parcial: 05025).

        Parâmetros
        ----------
            search_crit (str): Critério de busca
        """
        try:
            self.get(const.CRITERIA["URL"][search_crit])
        except:
            print("Ocorreu algum erro...")
        sleep(2)

    def new_search_summary(self, idx, row, search_msg):
        """Método que retorna uma nova linha para a página de resumo
        
        Parâmetros
        ----------
            idx (int): Linha Excel Entrada 
            
            row (list): Linha de busca. Tuple/List com 3 valores(Nome, CEP, Critério de Busca).

            search_msg (str): Mensagem de status da pesquisa

        Retorna
        ----------
            Dataframe com informações para a página de resumo"""
        search_name, search_CEP, search_crit = row

        new_search_summary = pd.DataFrame(
            [[idx, search_name, search_CEP, search_crit, const.MSG_STATUS[search_msg]]],
            columns=const.SUMMARY_HEADER,
        )

        return new_search_summary

    def paste_keys(self, text, elem):
        """Cola a informação no campo de busca.

        Parameters
        ----------
            text (str): Termo a ser pesquisado

            elem (WebElement): Elemento HTML do campo de busca
        """
        os.system("echo %s| clip" % text.strip())
        elem.send_keys(Keys.CONTROL, "v")

    def validate_CEP(self, search_param, no_crit=False):
        """Valida o valor do campo 'CEP'.

        Parâmetros
        ----------
            search_param (str): Parâmetro de busca do campo 'CEP'.
            
            no_crit (bool): Flag para as pesquisas que estão
            com o critério de busca inválido.

        Retorna
        ----------
            String com código de sucesso ou fracasso (com motivo).
            Os valores possíveis de retorno são as chaves da constante 'MSG_STATUS'.
        """
        # CEP vazio
        if not search_param:
            return "Denied_no_CEP" if not no_crit else "Nome"
        else:
            cep_complete = re.search(const.REGEX_CEP_COMPLETE, search_param)
            cep_incomplete = re.search(const.REGEX_CEP_INCOMPLETE, search_param)

        if cep_complete or cep_incomplete:
            new_param = f"0{search_param}" if len(search_param) == 4 else search_param
            has_results = self.execute_search("CEP", new_param)

            if has_results:
                if cep_complete:
                    return f'Success{"_no_criteria" if no_crit else ""}_CEP'
                else:
                    return f'Success{"_no_criteria" if no_crit else ""}_CEP_incomplete'

        return "Denied_invalid_CEP" if not no_crit else "Nome"

    def validate_demand(self, wb):
        """Valida a planilha a ser utilizada na pesquisa ( deve ter 3 colunas e pelo menos 1 linha de dados para a pesquisa).
        
        Parâmetros
        ----------
            wb (Workbook): Excel a ser validado.

        Retorna
        ----------
            Tabela de demanda caso a demanda seja validada.
            Lista vazia caso contrário.
        """
        ws = wb.active
        validation = (ws.max_column == 3) and (ws.max_row >= 1)
        search_table = list(ws.values) if validation else []
        return search_table

    def validate_name(self, search_param, no_crit=False):
        """Valida o valor do campo Nome.

        Parâmetros
        ----------
            search_param (str): Parâmetro de busca do campo 'Nome'.
            
            no_crit (bool): Flag para as pesquisas que estão
            com o critério de busca inválido.

        Retorna
        ----------
            String com código de sucesso ou fracasso (com motivo).
            Os valores possíveis de retorno são as chaves da constante 'MSG_STATUS'.
        """
        if not search_param:
            return f"Denied_no_name"

        elif re.search(const.REGEX_NAME, search_param):
            has_results = self.execute_search("Nome", search_param)

            if has_results:
                return f'Success{"_no_criteria" if no_crit else ""}_name'

        return f"Denied_invalid_name"

    def validate_search(self, row):
        """Valida a linha de entrada do Excel .
        
        Parâmetros
        ----------
            row: Linha de busca. Tuple/List com 3 valores(Nome, CEP, Critério de Busca).

        Retorna
        ----------
            Tuple com 3 valores: 
            1) Código de sucesso ou fracasso (com motivo)
            2) Critério de busca (ou `''` se fracasso).
            3) Parâmetro de busca (ou `''` se fracasso).
        """
        #  Nome,        CEP,     Critério de busca
        search_name, search_CEP, search_crit = row

        if search_crit == "CEP":
            return (self.validate_CEP(search_CEP), search_crit, search_CEP)
        elif search_crit == "Nome":
            return (self.validate_name(search_name), search_crit, search_name)
        else:
            # Primeiro tenta validar o CEP
            CEP_validation = self.validate_CEP(search_CEP, no_crit=1)

            if CEP_validation[0] == "S":
                # se der boa a validação do CEP
                return (CEP_validation, "CEP", search_CEP)

            # Se não der boa a validação do CEP, tenta validar o Nome
            name_validation = self.validate_name(search_name, no_crit=1)
            if name_validation[0] == "S":
                return (name_validation, "Nome", search_name)
            else:
                return (f"Denied_poor_parameters", "", "")
